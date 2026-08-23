import asyncio
from playwright.async_api import async_playwright
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd

async def scrape_full_catalog():
    print("Starting Full Market Intel Engine...")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page()

        url = "https://books.toscrape.com/"
        print(f"Coneccting to: {url}")

        await page.goto(url)
        await page.wait_for_load_state("networkidle")

        data = []

        loop = 1

        while loop <= 50:

            html = await page.content()
            sopa = BeautifulSoup(html, "html.parser")
            
            books = sopa.find_all("article", class_="product_pod")
            print(f"Conxion sucscesful, conected to {len(books)} elemets.")

            stars_convertor = {
                                "One": 1,
                                "Two": 2,
                                "Three": 3,
                                "Four": 4,
                                "Five": 5
                            }

            for i, book in enumerate(books, 1):
                enlace = book.select_one('h3 > a')
                titulo = enlace.get('title', 'No disponible') if enlace else "No disponible"

                precio_contenedor = book.select_one('.price_color')
                if precio_contenedor:
                    price = float(precio_contenedor.text.strip()[1:])
                else:
                    price = 0.00

                ratings = book.find('p', class_='star-rating')

                if ratings:
                    clases = ratings.get('class')
                    stars = clases[1]
                    stars = stars_convertor.get(stars, 0)

                else:
                    stars = 0

                book_numb = ((loop - 1) * 20) + i
                print(f"{book_numb}. {titulo} | {price} | {stars}/5")

                data.append({
                    "LIBRO": titulo,
                    "PRECIO": price,
                    "VALORACION": stars
                })

            boton_next = await page.query_selector("li.next a")

            if boton_next:
                await page.click("li.next a")

                await page.wait_for_load_state("networkidle")

                await asyncio.sleep(0.1) 

                loop += 1
                print(f"Next page {loop}/50")

            else:
                break
            
        print("Ended")

        excel_table = pd.DataFrame(data)
        excel_name = "books.xlsx"

        with pd.ExcelWriter(excel_name, engine='openpyxl') as writer:
            excel_table.to_excel(writer, index=False, sheet_name='Market Data')
            worksheet = writer.sheets['Market Data']

            fuente_negrita = Font(name="Calibri", size=16, bold=True)
            for cell in worksheet[1]: 
                cell.font = fuente_negrita
        
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)

                max_len = 0
                for row in range(1, worksheet.max_row + 1):
                    cell_value = worksheet.cell(row=row, column=col_idx).value
                    if cell_value:
                        max_len = max(max_len, len(str(cell_value)))
                
                worksheet.column_dimensions[col_letter].width = max(max_len + 5, 15)

if __name__ == "__main__":
    asyncio.run(scrape_full_catalog())